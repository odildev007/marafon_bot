from aiogram.types import Message, FSInputFile
import openpyxl
from openpyxl.styles import Font, Color, PatternFill, Border, Side
import sqlPrompts

async def send_reports_answer(message: Message):
    users = sqlPrompts.get_users()
    reyting = []
    for user in users:
        if user['status'] == "active":
            referals = sqlPrompts.get_user_referals(user["user_id"])
            ball = 0
            if referals:
                for referal in referals:
                    if referal['status'] == 'active':
                        ball += 1
            reyting.append((user["user_id"], user['name'], ball))
    reyting.sort(key=lambda x: x[2], reverse=True)

    if reyting:
        wb = openpyxl.Workbook()
        ws = wb.active
            
        ws.append(("ID", "Ism-familya", "Ball"))
        
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 5

        border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

        for i in "ABC": 
            ws[f'{i}1'].font = Font(color="FFFFFF", bold=True)  # Matn rangini qizil va qalin qilib sozlaymiz
            ws[f'{i}1'].fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")  # Orqa fonni yashil rangga sozlaymiz
        k = 1
        for i in reyting:
                ws.append((
                    str(i[0]),
                    i[1],
                    str(i[2])
                    ))
                k += 1
                ws[f"A{k}"].border = border
                ws[f"B{k}"].border = border
                ws[f"C{k}"].border = border


        wb.save("reports.xlsx")
        await message.answer_document(FSInputFile("reports.xlsx"))


async def get_db(message: Message):
    await message.answer_document(FSInputFile("data.db"))
